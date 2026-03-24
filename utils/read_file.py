# author=@mengmeng
# describe: 通用 Excel 用例读取器

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ExcelReader:
    """通用 Excel 测试用例读取器。

    将 Excel 文件中的每一行解析为 TestCase 字典，支持：
    - JSON 字段自动解析（headers/params/body）
    - run_status 过滤（FALSE 跳过）
    - expected_ 前缀列自动收集为断言期望值
    - 必填列校验

    Attributes:
        RESERVED_COLUMNS: 保留列名集合，不作为业务参数。
        JSON_COLUMNS: 需要 JSON 解析的列名集合。
        REQUIRED_COLUMNS: 必填列名集合。
    """

    RESERVED_COLUMNS: set[str] = {
        "case_id", "title", "method", "url",
        "headers", "params", "body", "type",
        "run_status", "is_run",
    }
    JSON_COLUMNS: set[str] = {"headers", "params", "body"}
    REQUIRED_COLUMNS: set[str] = {"method", "url"}

    def __init__(self, file_path: str, sheet_name: str = "Sheet1") -> None:
        """初始化 ExcelReader。

        Args:
            file_path: Excel 文件路径。
            sheet_name: Sheet 名称，默认 Sheet1。

        Raises:
            FileNotFoundError: 文件路径不存在时抛出。
            KeyError: Sheet 名称不存在时抛出。
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Sheet '{sheet_name}' 不存在，可用 Sheet: {wb.sheetnames}"
            )

        self._ws = wb[sheet_name]
        self._file_path = file_path

    def _validate_columns(self, headers: list[str]) -> None:
        """校验必填列是否存在。

        Args:
            headers: Excel 第一行列名列表。

        Raises:
            ValueError: 缺少必填列时抛出。
        """
        missing = self.REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"Excel 缺少必填列: {sorted(missing)}")

    def _parse_json_field(
        self, value: Any, row: int, col: str
    ) -> dict | None:
        """解析 JSON 字符串字段。

        Args:
            value: 单元格原始值。
            row: 行号（用于错误提示）。
            col: 列名（用于错误提示）。

        Returns:
            解析后的字典，或 None（值为空时）。

        Raises:
            ValueError: 值非空但无法解析为合法 JSON 时抛出。
        """
        if value is None or str(value).strip() == "":
            return None
        try:
            result = json.loads(str(value))
            if not isinstance(result, dict):
                raise ValueError
            return result
        except (json.JSONDecodeError, ValueError):
            raise ValueError(f"第{row}行 {col} 列 JSON 解析失败: {value}")

    def _parse_run_status(self, value: Any) -> bool:
        """解析 run_status 列值为布尔值。

        Args:
            value: 单元格原始值。

        Returns:
            仅当值大小写不敏感等于 "TRUE" 时返回 True，其余返回 False。
        """
        if value is None:
            return False
        return str(value).strip().upper() == "TRUE"

    def load_cases(self) -> list[dict]:
        """加载并解析所有测试用例。

        Returns:
            TestCase 字典列表，每条结构如下：
            {
                "case_id": str,
                "title": str,
                "method": str,
                "url": str,
                "headers": dict | None,
                "params": dict | None,
                "body": dict | None,
                "type": str | None,
                "run_status": bool,
                "expected": {"code": "200", "msg": "成功", ...}
            }

        Raises:
            ValueError: 必填列缺失或值为空时抛出。
        """
        rows = list(self._ws.iter_rows(values_only=True))
        if not rows:
            return []

        raw_headers: list[str] = [
            str(h) if h is not None else "" for h in rows[0]
        ]
        self._validate_columns(raw_headers)

        cases: list[dict] = []

        for row_idx, row in enumerate(rows[1:], start=2):
            row_data: dict[str, Any] = dict(zip(raw_headers, row))

            # run_status=FALSE 时跳过
            if not self._parse_run_status(row_data.get("run_status")):
                continue

            # 校验必填列
            for col in self.REQUIRED_COLUMNS:
                val = row_data.get(col)
                if val is None or str(val).strip() == "":
                    raise ValueError(f"第{row_idx}行 {col} 列不能为空")

            # 收集 expected_ 前缀列
            expected: dict[str, Any] = {}
            for key, val in row_data.items():
                if key.startswith("expected_"):
                    field = key[len("expected_"):]
                    expected[field] = val

            # 解析 JSON 字段
            headers = self._parse_json_field(
                row_data.get("headers"), row_idx, "headers"
            )
            params = self._parse_json_field(
                row_data.get("params"), row_idx, "params"
            )
            body = self._parse_json_field(
                row_data.get("body"), row_idx, "body"
            )

            case: dict = {
                "case_id": str(row_data.get("case_id", "")),
                "title": str(row_data.get("title", "")),
                "method": str(row_data.get("method", "")).upper(),
                "url": str(row_data.get("url", "")),
                "headers": headers,
                "params": params,
                "body": body,
                "type": str(row_data.get("type", "")) or None,
                "run_status": self._parse_run_status(row_data.get("run_status")),
                "expected": expected,
            }
            cases.append(case)

        return cases
