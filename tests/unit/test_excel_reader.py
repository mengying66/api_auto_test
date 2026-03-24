# describe: ExcelReader 单元测试

import json
import pytest
from openpyxl import Workbook
from pathlib import Path

from utils.read_file import ExcelReader


# ── 辅助函数 ──────────────────────────────────────────────────────────────────

def make_excel(tmp_path: Path, rows: list[list], sheet_name: str = "Sheet1") -> str:
    """在临时目录创建测试用 Excel 文件。

    Args:
        tmp_path: pytest 临时目录。
        rows: 行数据列表，第一行为表头。
        sheet_name: Sheet 名称。

    Returns:
        Excel 文件路径字符串。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    path = tmp_path / "test_cases.xlsx"
    wb.save(str(path))
    return str(path)


# ── 正常解析 ──────────────────────────────────────────────────────────────────

class TestLoadCasesNormal:
    """正常用例解析测试。"""

    def test_basic_parse(self, tmp_path: Path) -> None:
        """测试基本字段解析正确。"""
        rows = [
            ["case_id", "title", "method", "url", "headers", "params", "body", "type", "run_status", "expected_code", "expected_msg"],
            [1, "登录成功", "POST", "/api/login", '{"Content-Type":"application/json"}', None, '{"username":"admin","password":"123"}', "json", "TRUE", 200, "成功"],
        ]
        path = make_excel(tmp_path, rows)
        cases = ExcelReader(path).load_cases()

        assert len(cases) == 1
        c = cases[0]
        assert c["method"] == "POST"
        assert c["url"] == "/api/login"
        assert c["headers"] == {"Content-Type": "application/json"}
        assert c["body"] == {"username": "admin", "password": "123"}
        assert c["run_status"] is True
        assert c["expected"] == {"code": 200, "msg": "成功"}

    def test_expected_prefix_collected(self, tmp_path: Path) -> None:
        """测试 expected_ 前缀列被正确收集到 expected 子字典。"""
        rows = [
            ["case_id", "title", "method", "url", "run_status", "expected_code", "expected_msg", "expected_data"],
            [1, "测试", "GET", "/api/test", "TRUE", 200, "ok", "some_data"],
        ]
        path = make_excel(tmp_path, rows)
        cases = ExcelReader(path).load_cases()

        assert cases[0]["expected"] == {"code": 200, "msg": "ok", "data": "some_data"}

    def test_run_status_false_skipped(self, tmp_path: Path) -> None:
        """测试 run_status=FALSE 的行被跳过。"""
        rows = [
            ["case_id", "title", "method", "url", "run_status", "expected_code"],
            [1, "执行", "GET", "/api/a", "TRUE", 200],
            [2, "跳过", "GET", "/api/b", "FALSE", 200],
            [3, "也跳过", "GET", "/api/c", None, 200],
        ]
        path = make_excel(tmp_path, rows)
        cases = ExcelReader(path).load_cases()

        assert len(cases) == 1
        assert cases[0]["url"] == "/api/a"

    def test_null_json_fields_return_none(self, tmp_path: Path) -> None:
        """测试 headers/params/body 为空时返回 None。"""
        rows = [
            ["case_id", "title", "method", "url", "headers", "params", "body", "run_status"],
            [1, "空字段", "GET", "/api/test", None, None, None, "TRUE"],
        ]
        path = make_excel(tmp_path, rows)
        cases = ExcelReader(path).load_cases()

        c = cases[0]
        assert c["headers"] is None
        assert c["params"] is None
        assert c["body"] is None

    def test_run_status_case_insensitive(self, tmp_path: Path) -> None:
        """测试 run_status 大小写不敏感。"""
        rows = [
            ["case_id", "title", "method", "url", "run_status"],
            [1, "小写true", "GET", "/api/a", "true"],
            [2, "混合True", "GET", "/api/b", "True"],
        ]
        path = make_excel(tmp_path, rows)
        cases = ExcelReader(path).load_cases()

        assert len(cases) == 2

    def test_multiple_sheets_independent(self, tmp_path: Path) -> None:
        """测试多 Sheet 独立解析互不干扰。"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["case_id", "title", "method", "url", "run_status"])
        ws1.append([1, "Sheet1用例", "GET", "/api/s1", "TRUE"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["case_id", "title", "method", "url", "run_status"])
        ws2.append([2, "Sheet2用例", "POST", "/api/s2", "TRUE"])

        path = tmp_path / "multi.xlsx"
        wb.save(str(path))

        cases1 = ExcelReader(str(path), "Sheet1").load_cases()
        cases2 = ExcelReader(str(path), "Sheet2").load_cases()

        assert len(cases1) == 1 and cases1[0]["url"] == "/api/s1"
        assert len(cases2) == 1 and cases2[0]["url"] == "/api/s2"


# ── 错误场景 ──────────────────────────────────────────────────────────────────

class TestLoadCasesErrors:
    """错误场景测试。"""

    def test_file_not_found(self) -> None:
        """测试文件不存在时抛出 FileNotFoundError，且消息包含路径。"""
        bad_path = "/nonexistent/path/cases.xlsx"
        with pytest.raises(FileNotFoundError, match=bad_path):
            ExcelReader(bad_path)

    def test_sheet_not_found(self, tmp_path: Path) -> None:
        """测试 Sheet 不存在时抛出 KeyError。"""
        rows = [["case_id", "method", "url", "run_status"]]
        path = make_excel(tmp_path, rows)
        with pytest.raises(KeyError, match="NoSuchSheet"):
            ExcelReader(path, "NoSuchSheet")

    def test_missing_required_column(self, tmp_path: Path) -> None:
        """测试缺少必填列 method 时抛出 ValueError。"""
        rows = [
            ["case_id", "title", "url", "run_status"],  # 缺少 method
            [1, "测试", "/api/test", "TRUE"],
        ]
        path = make_excel(tmp_path, rows)
        with pytest.raises(ValueError, match="method"):
            ExcelReader(path).load_cases()

    def test_required_column_empty_value(self, tmp_path: Path) -> None:
        """测试必填列 url 值为空时抛出 ValueError，且消息包含行号和列名。"""
        rows = [
            ["case_id", "title", "method", "url", "run_status"],
            [1, "测试", "GET", None, "TRUE"],  # url 为空
        ]
        path = make_excel(tmp_path, rows)
        with pytest.raises(ValueError, match="url"):
            ExcelReader(path).load_cases()

    def test_invalid_json_in_headers(self, tmp_path: Path) -> None:
        """测试 headers 列非法 JSON 时抛出 ValueError，包含行号和列名。"""
        rows = [
            ["case_id", "title", "method", "url", "headers", "run_status"],
            [1, "测试", "GET", "/api/test", "not-json", "TRUE"],
        ]
        path = make_excel(tmp_path, rows)
        with pytest.raises(ValueError, match="headers"):
            ExcelReader(path).load_cases()

    def test_invalid_json_in_body(self, tmp_path: Path) -> None:
        """测试 body 列非法 JSON 时抛出 ValueError，包含行号和列名。"""
        rows = [
            ["case_id", "title", "method", "url", "body", "run_status"],
            [1, "测试", "POST", "/api/test", "{bad json}", "TRUE"],
        ]
        path = make_excel(tmp_path, rows)
        with pytest.raises(ValueError, match="body"):
            ExcelReader(path).load_cases()

    def test_invalid_json_in_params(self, tmp_path: Path) -> None:
        """测试 params 列非法 JSON 时抛出 ValueError，包含行号和列名。"""
        rows = [
            ["case_id", "title", "method", "url", "params", "run_status"],
            [1, "测试", "GET", "/api/test", "invalid", "TRUE"],
        ]
        path = make_excel(tmp_path, rows)
        with pytest.raises(ValueError, match="params"):
            ExcelReader(path).load_cases()
