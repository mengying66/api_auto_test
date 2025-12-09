# author=@mengmeng
# time：2025/12/5 16：49：47
# describe：读取excel文件来执行接口测试

import json
import time
import requests
import pytest
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Dict
import subprocess


# -------------------------- 1. Key持久化工具类 --------------------------
class AuthKeyManager:
    """登录Key持久化管理：保存/读取/校验过期"""

    def __init__(self, key_file: str = "auth_key.json"):
        self.key_file = Path(key_file)
        # 初始化文件（若不存在则创建空文件）
        if not self.key_file.exists():
            with open(self.key_file, "w", encoding="utf-8") as f:
                json.dump({"token": "", "expire_time": 0}, f)

    def save_key(self, token: str, expire_seconds: int = 3600):
        """
        保存Key到本地
        :param token: 登录获取的Key/Token
        :param expire_seconds: 过期时间（默认1小时）
        """
        auth_data = {
            "token": token,
            "expire_time": time.time() + expire_seconds  # 过期时间戳
        }
        with open(self.key_file, "w", encoding="utf-8") as f:
            json.dump(auth_data, f, ensure_ascii=False, indent=2)
        print(f"Key已保存到本地：{self.key_file}")

    def get_valid_key(self) -> str:
        """获取有效Key：未过期则返回，过期返回空"""
        with open(self.key_file, "r", encoding="utf-8") as f:
            auth_data = json.load(f)

        # 校验Key是否有效（存在+未过期）
        if auth_data.get("token") and time.time() < auth_data.get("expire_time", 0):
            print("使用本地缓存的有效Key")
            return auth_data["token"]
        else:
            print("本地Key已过期/不存在")
            return ""

    def clear_key(self):
        """清空本地Key（用于重新登录）"""
        self.save_key("", 0)


# -------------------------- 2. Excel读取工具（适配你的表头） --------------------------
class ExcelReader:
    def __init__(self, file_path: str, sheet_name: str = "Sheet1"):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel文件不存在：{self.file_path}")
        self.wb = load_workbook(self.file_path, data_only=True)
        self.ws = self.wb[sheet_name]
        self.headers = self._get_headers()

    def _get_headers(self) -> List[str]:
        """获取Excel表头（第一行）"""
        return [cell.value.strip() if cell.value else "" for cell in self.ws[1]]

    def read_test_cases(self) -> List[Dict]:
        """读取所有测试用例，返回字典列表"""
        cases = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            case = dict(zip(self.headers, row))
            if not case.get("case_id"):
                continue
            # 空值处理
            for key, value in case.items():
                case[key] = value if value is not None else ""
            cases.append(case)
        self.wb.close()
        return cases


# -------------------------- 3. 核心请求类（集成Key管理） --------------------------
class ApiTestClient:
    def __init__(self, base_url: str, login_case: Dict = None):
        self.base_url = base_url
        self.session = requests.Session()
        self.key_manager = AuthKeyManager()  # 初始化Key管理器
        self.login_case = login_case  # 登录用例（用于自动重新登录）

    def login_and_save_key(self) -> str:
        """执行登录用例，获取Key并持久化"""
        if not self.login_case:
            raise ValueError("未配置登录用例，无法获取Key")

        # 解析登录用例参数
        login_url = self.base_url + self.login_case["url"]
        login_method = self.login_case["method"].upper()
        login_headers = json.loads(self.login_case["headers"]) if self.login_case["headers"] else {}
        login_body = json.loads(self.login_case["body"]) if self.login_case["body"] else {}

        # 发送登录请求
        print(f"执行登录接口：{self.login_case['title']}")
        if login_method == "POST":
            response = self.session.post(login_url, headers=login_headers, json=login_body, timeout=10)
        else:
            response = self.session.request(login_method, login_url, headers=login_headers, params=login_body,
                                            timeout=10)

        # 解析Key（根据你的接口返回格式调整，示例：response.json()["data"]["token"]）
        response_json = response.json()
        if response.status_code != 200 or "token" not in response_json.get("data", {}):
            raise RuntimeError(f"登录失败：{response_json}")

        token = response_json["data"]["token"]
        # 保存Key（设置过期时间，示例：2小时）
        self.key_manager.save_key(token, expire_seconds=7200)
        return token

    def get_auth_key(self) -> str:
        """获取有效Key：优先读本地，无效则重新登录"""
        key = self.key_manager.get_valid_key()
        if not key:
            # 本地无有效Key，执行登录并获取新Key
            key = self.login_and_save_key()
        return key

    def send_request(self, case: Dict) -> requests.Response:
        """发送接口请求（自动注入鉴权Key）"""
        # 1. 基础参数解析
        method = case["method"].upper()
        url = self.base_url + case["url"]
        params = json.loads(case["params"]) if case["params"] else {}
        headers = json.loads(case["headers"]) if case["headers"] else {}
        body = json.loads(case["body"]) if case["body"] and case["type"] == "json" else case["body"]

        # 2. 注入鉴权Key（根据你的接口鉴权方式调整，示例：Bearer Token）
        auth_key = self.get_auth_key()
        headers["Authorization"] = f"Bearer {auth_key}"  # 适配你的鉴权头格式

        # 3. 发送请求
        try:
            if method == "GET":
                response = self.session.get(url, params=params, headers=headers, timeout=10)
            elif method == "POST":
                if case["type"] == "json":
                    response = self.session.post(url, params=params, headers=headers, json=body, timeout=10)
                else:  # form表单
                    response = self.session.post(url, params=params, headers=headers, data=body, timeout=10)
            elif method == "PUT":
                response = self.session.put(url, params=params, headers=headers, json=body, timeout=10)
            elif method == "DELETE":
                response = self.session.delete(url, params=params, headers=headers, timeout=10)
            else:
                raise ValueError(f"不支持的请求方法：{method}")
            return response
        except Exception as e:
            raise RuntimeError(f"接口请求失败【{case['case_id']}-{case['title']}】：{str(e)}")


# -------------------------- 4. 测试执行逻辑 --------------------------
# 配置项（根据你的项目修改）
BASE_URL = "https://test-admin-allreaday.com"  # 你的接口基础URL
EXCEL_PATH = "autotest_case1.xlsx"  # 你的Excel用例路径


# 从Excel中筛选登录用例（标记is_login_case=True）
def get_login_case(all_cases: List[Dict]) -> Dict:
    login_cases = [case for case in all_cases if case.get("is_login_case", "").lower() == "true"]
    if not login_cases:
        raise ValueError("Excel中未找到标记is_login_case=True的登录用例")
    return login_cases[0]


# 初始化
excel_reader = ExcelReader(EXCEL_PATH)
all_test_cases = excel_reader.read_test_cases()
login_case = get_login_case(all_test_cases)
api_client = ApiTestClient(base_url=BASE_URL, login_case=login_case)

# 过滤掉登录用例（避免重复执行）
business_cases = [case for case in all_test_cases if case.get("is_login_case", "").lower() != "true"]


# -------------------------- 5. 测试用例执行 --------------------------
@pytest.mark.parametrize("case", business_cases)
def test_api_auto(case):
    """参数化执行业务接口测试"""
    print(f"\n===== 执行用例：[{case['case_id']}] {case['title']} =====")

    # 发送请求
    response = api_client.send_request(case)

    # 结果断言（适配你的Excel表头）
    try:
        # 断言状态码
        assert response.status_code == int(case["expected_code"]), \
            f"状态码断言失败：预期{case['expected_code']}，实际{response.status_code}"

        # 断言响应消息
        resp_json = response.json()
        assert case["expected_msg"] in str(resp_json), \
            f"响应消息断言失败：预期包含'{case['expected_msg']}'，实际{resp_json}"

        print(f"用例[{case['case_id']}]执行成功！")
    except json.JSONDecodeError:
        # 非JSON响应断言
        assert case["expected_msg"] in response.text, \
            f"响应消息断言失败：预期包含'{case['expected_msg']}'，实际{response.text}"
    except AssertionError as e:
        print(f"用例[{case['case_id']}]执行失败：{str(e)}")
        raise


# -------------------------- 6. 执行入口 --------------------------
if __name__ == "__main__":
    # 执行测试并生成Allure报告（可选）
    pytest.main([
        __file__,
        "-vs",
        "--alluredir=./reports/allure-results"
    ])
    # 可选：自动生成Allure HTML报告
    subprocess.run(["allure", "generate", "./reports/allure-results", "-o", "./reports/allure-report", "--clean"])